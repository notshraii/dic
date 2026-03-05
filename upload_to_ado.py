#!/usr/bin/env python3
"""Upload test cases from CSV or Excel (.xlsx) to Azure DevOps (ADO).

Reads a CSV or Excel file containing test case definitions and creates
Test Case work items in Azure DevOps via the REST API, including
structured test steps.

Supports CSV and Excel (.xlsx/.xls) input. Two layouts are auto-detected:

  FORMAT A  - One row per test case (flat):
    Columns: ID, Title, Area, Priority, Description, Prerequisites,
             TestSteps, ExpectedResult, Notes
    TestSteps are newline-separated numbered lines within one cell.

  FORMAT B  - One row per test step (expanded):
    Columns: [ID,] Work Item Type, Title, [Priority, [Area Path,]]
             Test Step, Step Action, Step Expected
    Rows with the same Title are grouped into a single test case.
    Continuation rows may leave Title (and other leading columns) blank.

The script auto-detects the format from the header row.

Usage:
    python upload_to_ado.py <file> [options]

Examples:
    # Dry-run from Excel (no API calls, prints what would be created)
    python upload_to_ado.py my_tests.xlsx --dry-run

    # Excel with a specific sheet name
    python upload_to_ado.py my_tests.xlsx --sheet "Regression Tests" --dry-run

    # Dry-run from CSV
    python upload_to_ado.py ADO_TEST_CASES_IMPORT_FIXED.csv --dry-run

    # Upload to ADO
    python upload_to_ado.py ADO_TEST_CASES_IMPORT_FIXED.csv

    # Specify project and area path
    python upload_to_ado.py ADO_TEST_CASES.csv --project MyProject --area "MyProject\\Testing"

    # Override org/PAT via CLI instead of .env
    python upload_to_ado.py tests.csv --org https://dev.azure.com/myorg --pat MY_TOKEN

Environment variables (or .env file):
    ADO_ORG_URL          - Azure DevOps organization URL
    ADO_PROJECT          - Project name
    ADO_PAT              - Personal Access Token
    ADO_AREA_PATH        - Default area path for new test cases
    ADO_TEST_PLAN_ID     - (optional) Add created cases to this test plan
    ADO_TEST_SUITE_ID    - (optional) Add created cases to this test suite
"""

import argparse
import base64
import csv
import json
import os
import re
import sys
import time
from dataclasses import dataclass, field
from pathlib import Path

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

try:
    import requests
except ImportError:
    print("ERROR: 'requests' package is required. Install with: pip install requests")
    sys.exit(1)


# ---------------------------------------------------------------------------
# Data model
# ---------------------------------------------------------------------------

@dataclass
class TestStep:
    index: int
    action: str
    expected: str


@dataclass
class TestCase:
    title: str
    steps: list[TestStep] = field(default_factory=list)
    priority: int = 2
    area_path: str = ""
    description: str = ""
    prerequisites: str = ""
    expected_result: str = ""
    notes: str = ""
    csv_id: str = ""


# ---------------------------------------------------------------------------
# CSV parsing
# ---------------------------------------------------------------------------

def _normalise_header(h: str) -> str:
    return re.sub(r"[^a-z0-9]", "", h.strip().lower())


def detect_format(headers: list[str]) -> str:
    """Return 'A' for flat format, 'B' for step-per-row format."""
    norm = [_normalise_header(h) for h in headers]
    if "stepaction" in norm:
        return "B"
    if "teststeps" in norm or "description" in norm:
        return "A"
    raise ValueError(
        f"Cannot detect CSV format from headers: {headers}\n"
        "Expected either 'Step Action' (format B) or 'TestSteps'/'Description' (format A)."
    )


def _parse_steps_text(text: str) -> list[TestStep]:
    """Parse numbered step text like '1. Do X\\n2. Do Y' into TestStep list."""
    steps = []
    for line in text.strip().splitlines():
        line = line.strip()
        if not line:
            continue
        m = re.match(r"^(\d+)\.\s*(.*)", line)
        if m:
            steps.append(TestStep(index=int(m.group(1)), action=m.group(2), expected=""))
        else:
            steps.append(TestStep(index=len(steps) + 1, action=line, expected=""))
    return steps


def parse_format_a(reader, headers: list[str]) -> list[TestCase]:
    """Parse flat CSV: one row per test case."""
    norm = {_normalise_header(h): i for i, h in enumerate(headers)}

    def col(row, key, default=""):
        idx = norm.get(key)
        if idx is None or idx >= len(row):
            return default
        return row[idx].strip()

    cases = []
    for row in reader:
        if not any(c.strip() for c in row):
            continue
        title = col(row, "title")
        if not title:
            continue

        tc = TestCase(
            title=title,
            csv_id=col(row, "id"),
            priority=int(col(row, "priority", "2") or "2"),
            area_path=col(row, "area") or col(row, "areapath"),
            description=col(row, "description"),
            prerequisites=col(row, "prerequisites"),
            expected_result=col(row, "expectedresult"),
            notes=col(row, "notes"),
        )

        steps_text = col(row, "teststeps")
        if steps_text:
            tc.steps = _parse_steps_text(steps_text)

        cases.append(tc)
    return cases


def parse_format_b(reader, headers: list[str]) -> list[TestCase]:
    """Parse expanded CSV: one row per test step, grouped by title."""
    norm = {_normalise_header(h): i for i, h in enumerate(headers)}

    def col(row, key, default=""):
        idx = norm.get(key)
        if idx is None or idx >= len(row):
            return default
        return row[idx].strip()

    cases: list[TestCase] = []
    current: TestCase | None = None

    for row in reader:
        if not any(c.strip() for c in row):
            continue

        title = col(row, "title")
        step_action = col(row, "stepaction")
        step_expected = col(row, "stepexpected")
        step_num_str = col(row, "teststep") or col(row, "stepnumber")

        is_new_case = bool(title) and (current is None or title != current.title)

        if is_new_case:
            if current is not None:
                cases.append(current)
            current = TestCase(
                title=title,
                csv_id=col(row, "id"),
                priority=int(col(row, "priority", "2") or "2"),
                area_path=col(row, "areapath") or col(row, "area"),
            )

        if current is None:
            continue

        if step_action:
            try:
                step_idx = int(step_num_str) if step_num_str else len(current.steps) + 1
            except ValueError:
                step_idx = len(current.steps) + 1
            current.steps.append(TestStep(
                index=step_idx,
                action=step_action,
                expected=step_expected,
            ))

    if current is not None:
        cases.append(current)

    return cases


def _load_excel(path: Path, sheet_name: str | None = None) -> tuple[list[str], list[list[str]]]:
    """Read an Excel file and return (headers, rows) as string lists."""
    try:
        import openpyxl
    except ImportError:
        print("ERROR: 'openpyxl' package is required for .xlsx files. Install with: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    if sheet_name:
        if sheet_name not in wb.sheetnames:
            print(f"Available sheets: {wb.sheetnames}")
            raise ValueError(f"Sheet '{sheet_name}' not found in {path.name}")
        ws = wb[sheet_name]
    else:
        ws = wb.active
    print(f"Reading sheet: '{ws.title}'")

    rows_iter = ws.iter_rows(values_only=True)
    raw_headers = next(rows_iter)
    headers = [str(h).strip() if h is not None else "" for h in raw_headers]

    rows = []
    for raw_row in rows_iter:
        rows.append([str(c).strip() if c is not None else "" for c in raw_row])

    wb.close()
    return headers, rows


def load_test_cases(file_path: str, sheet_name: str | None = None) -> list[TestCase]:
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"File not found: {file_path}")

    is_excel = path.suffix.lower() in (".xlsx", ".xls")

    if is_excel:
        headers, rows = _load_excel(path, sheet_name)
    else:
        with open(path, newline="", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            headers = next(reader)
            rows = list(reader)

    fmt = detect_format(headers)
    label = "A (flat, one row per test case)" if fmt == "A" else "B (one row per step)"
    print(f"Detected format: {label}")

    if fmt == "A":
        return parse_format_a(iter(rows), headers)
    else:
        return parse_format_b(iter(rows), headers)


# ---------------------------------------------------------------------------
# ADO REST API client
# ---------------------------------------------------------------------------

def _xml_escape(text: str) -> str:
    return (
        text.replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;")
            .replace('"', "&quot;")
            .replace("'", "&apos;")
    )


def build_steps_xml(steps: list[TestStep]) -> str:
    """Build the XML blob ADO expects for Microsoft.VSTS.TCM.Steps."""
    if not steps:
        return ""
    parts = ['<steps id="0" last="{}">'.format(len(steps))]
    for s in steps:
        parts.append(
            '<step id="{}" type="ActionStep">'
            '<parameterizedString isformatted="true">{}</parameterizedString>'
            '<parameterizedString isformatted="true">{}</parameterizedString>'
            '</step>'.format(s.index, _xml_escape(s.action), _xml_escape(s.expected))
        )
    parts.append("</steps>")
    return "".join(parts)


class AdoClient:
    API_VERSION = "7.1"

    def __init__(self, org_url: str, project: str, pat: str):
        self.org_url = org_url.rstrip("/")
        self.project = project
        self.pat = pat
        self.session = requests.Session()
        token_bytes = base64.b64encode(f":{pat}".encode()).decode()
        self.session.headers.update({
            "Authorization": f"Basic {token_bytes}",
            "Content-Type": "application/json-patch+json",
        })

    def _api(self, path: str, api_version: str | None = None) -> str:
        ver = api_version or self.API_VERSION
        sep = "&" if "?" in path else "?"
        return f"{self.org_url}/{self.project}/_apis/{path}{sep}api-version={ver}"

    def create_test_case(self, tc: TestCase, area_path: str = "") -> dict:
        """Create a single Test Case work item and return the API response."""
        patch_doc = [
            {"op": "add", "path": "/fields/System.Title", "value": tc.title},
            {"op": "add", "path": "/fields/Microsoft.VSTS.Common.Priority", "value": tc.priority},
        ]

        effective_area = tc.area_path or area_path
        if effective_area:
            patch_doc.append({"op": "add", "path": "/fields/System.AreaPath", "value": effective_area})

        if tc.description:
            patch_doc.append({"op": "add", "path": "/fields/System.Description", "value": tc.description})

        if tc.steps:
            steps_xml = build_steps_xml(tc.steps)
            patch_doc.append({"op": "add", "path": "/fields/Microsoft.VSTS.TCM.Steps", "value": steps_xml})

        if tc.prerequisites:
            patch_doc.append({
                "op": "add",
                "path": "/fields/Microsoft.VSTS.TCM.Parameters",
                "value": tc.prerequisites,
            })

        url = self._api("wit/workitems/$Test Case")
        resp = self.session.post(url, json=patch_doc)
        resp.raise_for_status()
        return resp.json()

    def add_to_test_suite(self, plan_id: int, suite_id: int, case_ids: list[int]) -> dict:
        """Add test cases to a test suite within a test plan."""
        ids_str = ",".join(str(i) for i in case_ids)
        url = self._api(
            f"testplan/Plans/{plan_id}/Suites/{suite_id}/TestCase?testCaseIds={ids_str}",
            api_version="7.1",
        )
        self.session.headers["Content-Type"] = "application/json"
        resp = self.session.post(url, json=[])
        self.session.headers["Content-Type"] = "application/json-patch+json"
        resp.raise_for_status()
        return resp.json()


# ---------------------------------------------------------------------------
# Main upload logic
# ---------------------------------------------------------------------------

def upload_test_cases(
    file_path: str,
    org_url: str,
    project: str,
    pat: str,
    area_path: str = "",
    plan_id: int | None = None,
    suite_id: int | None = None,
    dry_run: bool = False,
    delay: float = 0.5,
    sheet_name: str | None = None,
) -> list[dict]:
    test_cases = load_test_cases(file_path, sheet_name=sheet_name)
    print(f"Loaded {len(test_cases)} test case(s) from {file_path}\n")

    if not test_cases:
        print("No test cases found. Check CSV file format.")
        return []

    if dry_run:
        print("=== DRY RUN (no API calls) ===\n")
        for i, tc in enumerate(test_cases, 1):
            print(f"  [{i}] {tc.title}")
            print(f"      Priority: {tc.priority}  |  Area: {tc.area_path or area_path or '(default)'}")
            if tc.description:
                print(f"      Description: {tc.description[:80]}...")
            print(f"      Steps: {len(tc.steps)}")
            for s in tc.steps:
                print(f"        {s.index}. {s.action}")
                if s.expected:
                    print(f"           => {s.expected}")
            print()
        print(f"Total: {len(test_cases)} test cases with {sum(len(tc.steps) for tc in test_cases)} steps.")
        return []

    client = AdoClient(org_url, project, pat)
    results = []
    created_ids = []

    for i, tc in enumerate(test_cases, 1):
        try:
            result = client.create_test_case(tc, area_path=area_path)
            work_item_id = result.get("id", "?")
            url = result.get("_links", {}).get("html", {}).get("href", "")
            print(f"  [{i}/{len(test_cases)}] Created: #{work_item_id} - {tc.title}")
            if url:
                print(f"      URL: {url}")
            results.append(result)
            created_ids.append(work_item_id)
        except requests.HTTPError as e:
            print(f"  [{i}/{len(test_cases)}] FAILED: {tc.title}")
            print(f"      Error: {e}")
            if e.response is not None:
                print(f"      Response: {e.response.text[:300]}")
            results.append({"error": str(e), "title": tc.title})

        if i < len(test_cases):
            time.sleep(delay)

    success_count = sum(1 for r in results if "id" in r)
    fail_count = len(results) - success_count
    print(f"\nDone: {success_count} created, {fail_count} failed out of {len(test_cases)} total.")

    if plan_id and suite_id and created_ids:
        valid_ids = [cid for cid in created_ids if isinstance(cid, int)]
        if valid_ids:
            print(f"\nAdding {len(valid_ids)} test case(s) to Plan {plan_id} / Suite {suite_id}...")
            try:
                client.add_to_test_suite(plan_id, suite_id, valid_ids)
                print("  Added successfully.")
            except requests.HTTPError as e:
                print(f"  Failed to add to suite: {e}")

    return results


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Upload test cases from CSV or Excel to Azure DevOps",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument("file", help="Path to CSV or Excel (.xlsx) file containing test cases")
    parser.add_argument("--sheet", default=None,
                        help="Excel sheet name to read (default: active sheet)")
    parser.add_argument("--org", default=os.getenv("ADO_ORG_URL", ""),
                        help="ADO org URL (or set ADO_ORG_URL env var)")
    parser.add_argument("--project", default=os.getenv("ADO_PROJECT", ""),
                        help="ADO project name (or set ADO_PROJECT env var)")
    parser.add_argument("--pat", default=os.getenv("ADO_PAT", ""),
                        help="Personal Access Token (or set ADO_PAT env var)")
    parser.add_argument("--area", default=os.getenv("ADO_AREA_PATH", ""),
                        help="Default area path for test cases")
    parser.add_argument("--plan-id", type=int, default=int(os.getenv("ADO_TEST_PLAN_ID", "0") or "0"),
                        help="Test Plan ID to add cases to")
    parser.add_argument("--suite-id", type=int, default=int(os.getenv("ADO_TEST_SUITE_ID", "0") or "0"),
                        help="Test Suite ID to add cases to")
    parser.add_argument("--dry-run", action="store_true",
                        help="Parse file and print what would be created, without calling the API")
    parser.add_argument("--delay", type=float, default=0.5,
                        help="Seconds to wait between API calls (default: 0.5)")

    args = parser.parse_args()

    if not args.dry_run:
        missing = []
        if not args.org:
            missing.append("--org / ADO_ORG_URL")
        if not args.project:
            missing.append("--project / ADO_PROJECT")
        if not args.pat:
            missing.append("--pat / ADO_PAT")
        if missing:
            parser.error(
                f"The following are required for upload (use --dry-run to skip):\n  "
                + "\n  ".join(missing)
            )

    upload_test_cases(
        file_path=args.file,
        org_url=args.org,
        project=args.project,
        pat=args.pat,
        area_path=args.area,
        plan_id=args.plan_id or None,
        suite_id=args.suite_id or None,
        dry_run=args.dry_run,
        delay=args.delay,
        sheet_name=args.sheet,
    )


if __name__ == "__main__":
    main()
